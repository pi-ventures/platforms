#!/usr/bin/env python3
"""
Import all data into DataHub PostgreSQL database.
Run after setup.bat has created the database and schema.

Usage:
    pip install psycopg2-binary openpyxl
    python import_data.py --password YOUR_POSTGRES_PASSWORD
"""

import csv
import argparse
import openpyxl
import psycopg2
import psycopg2.extras
import os
import time

def connect(password, host='localhost', port=5432, user='postgres', dbname='datahub'):
    return psycopg2.connect(host=host, port=port, user=user, password=password, dbname=dbname)

def import_villages(conn, village_file):
    """Import 5.96 lakh villages/towns from census CSV"""
    print("=" * 60)
    print("IMPORTING VILLAGES...")
    cur = conn.cursor()
    batch = []
    count = 0

    with open(village_file, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.DictReader(f)
        for row in reader:
            village_name = row.get('village_name', '').strip()
            town_name = row.get('town_name', '').strip()
            place_name = row.get('place_name', '').strip()
            name = village_name or town_name or place_name
            if not name:
                continue

            loc_type = 'TOWN' if town_name and not village_name else 'VILLAGE'
            batch.append((
                name, loc_type,
                row.get('subdistrict_name', '').strip() or None,
                row.get('district_name', '').strip() or None,
                row.get('state_name', '').strip() or None,
                row.get('LGD_code', '').strip() or None,
                row.get('shrid2', '').strip() or None,
                row.get('local_body_name', '').strip() or None,
                'India_Village_List_Census',
            ))
            count += 1

            if len(batch) >= 5000:
                psycopg2.extras.execute_values(cur,
                    """INSERT INTO locations (name, location_type, sub_district, district, state,
                       lgd_code, census_code_2011, post_office, source)
                       VALUES %s""",
                    batch, template="(%s, %s, %s, %s, %s, %s, %s, %s, %s)")
                conn.commit()
                print(f"  {count:,} imported...")
                batch = []

    if batch:
        psycopg2.extras.execute_values(cur,
            """INSERT INTO locations (name, location_type, sub_district, district, state,
               lgd_code, census_code_2011, post_office, source)
               VALUES %s""",
            batch, template="(%s, %s, %s, %s, %s, %s, %s, %s, %s)")
        conn.commit()

    print(f"  DONE: {count:,} villages/towns\n")
    cur.close()

def import_railways(conn, xlsx_file):
    """Import railway stations into company_locations"""
    print("=" * 60)
    print("IMPORTING RAILWAY STATIONS...")
    cur = conn.cursor()

    cur.execute("SELECT id FROM companies WHERE name='Ministry of Railways'")
    railways_id = cur.fetchone()[0]

    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    ws = wb['🚂 Railway Stations']

    def s(v): return str(v).strip() if v else None
    def f(v):
        try: return float(v) if v else None
        except: return None

    batch = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        batch.append((
            railways_id, s(row[2]), s(row[1]), 'RAILWAY_STATION',
            s(row[6]), s(row[5]), s(row[7]), f(row[8]), f(row[9]),
            s(row[4]), s(row[3]), 'KnowledgeHub_IndianRailways',
        ))

    psycopg2.extras.execute_values(cur,
        """INSERT INTO company_locations (company_id, branch_name, branch_code, branch_type,
           district, state, pincode, latitude, longitude, zone, zone_full, source)
           VALUES %s""",
        batch, template="(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
    conn.commit()
    print(f"  DONE: {len(batch):,} stations\n")
    wb.close()
    cur.close()

def import_post_offices(conn, xlsx_file):
    """Import post offices into company_locations"""
    print("=" * 60)
    print("IMPORTING POST OFFICES...")
    cur = conn.cursor()

    cur.execute("SELECT id FROM companies WHERE name='Department of Posts'")
    indiapost_id = cur.fetchone()[0]

    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    ws = wb['📮 PIN Cross-Reference']

    def s(v): return str(v).strip() if v else None

    batch = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        pincode = str(row[0]).strip()
        state, district = s(row[1]), s(row[2])
        all_pos = str(row[7]).strip() if row[7] else ''

        if all_pos and all_pos != 'None':
            for po_name in all_pos.split(' | '):
                po_name = po_name.strip()
                if po_name and '...' not in po_name and '+' not in po_name:
                    po_type = None
                    if 'H.O' in po_name or 'GPO' in po_name: po_type = 'HO'
                    elif 'S.O' in po_name: po_type = 'SO'
                    elif 'B.O' in po_name: po_type = 'BO'
                    clean = po_name.replace(' H.O','').replace(' S.O','').replace(' B.O','').replace(' GPO','').strip()
                    batch.append((indiapost_id, clean, pincode, 'POST_OFFICE', district, state, pincode, po_type, 'LogisticsNetwork_Unified'))

    psycopg2.extras.execute_values(cur,
        """INSERT INTO company_locations (company_id, branch_name, branch_code, branch_type,
           district, state, pincode, zone, source)
           VALUES %s""",
        batch, template="(%s, %s, %s, %s, %s, %s, %s, %s, %s)")
    conn.commit()
    print(f"  DONE: {len(batch):,} post offices\n")
    wb.close()
    cur.close()

def summary(conn):
    cur = conn.cursor()
    print("=" * 60)
    print("DATABASE SUMMARY")
    print("=" * 60)
    tables = ['people', 'companies', 'people_companies', 'contact_identifiers',
              'locations', 'people_locations', 'company_locations', 'relationships']
    total = 0
    for t in tables:
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        cnt = cur.fetchone()[0]
        total += cnt
        print(f"  {t:30s} {cnt:>10,} rows")
    print(f"\n  TOTAL: {total:,} rows")
    cur.close()

def main():
    parser = argparse.ArgumentParser(description='Import data into DataHub PostgreSQL')
    parser.add_argument('--password', required=True, help='PostgreSQL password')
    parser.add_argument('--host', default='localhost')
    parser.add_argument('--port', type=int, default=5432)
    parser.add_argument('--data-dir', default='.', help='Directory containing data files')
    args = parser.parse_args()

    conn = connect(args.password, args.host, args.port)
    print("Connected to PostgreSQL\n")

    data_dir = args.data_dir
    desktop = os.path.join(os.path.expanduser('~'), 'Desktop')

    # Import villages - check multiple locations
    village_file = os.path.join(data_dir, 'India Village list.csv')
    if not os.path.exists(village_file):
        village_file = os.path.join(desktop, 'PINs', 'India Village list.csv')
    if not os.path.exists(village_file):
        village_file = os.path.join(desktop, 'India Village list.csv')
    if os.path.exists(village_file):
        import_villages(conn, village_file)
    else:
        print(f"Skipping villages: not found in {data_dir} or Desktop\\PINs")

    # Import railways + post offices - check multiple locations
    xlsx_file = os.path.join(data_dir, 'LogisticsNetwork_IndianRailways_IndiaPost_Unified.xlsx')
    if not os.path.exists(xlsx_file):
        xlsx_file = os.path.join(desktop, 'LogisticsNetwork_IndianRailways_IndiaPost_Unified.xlsx')
    if os.path.exists(xlsx_file):
        import_railways(conn, xlsx_file)
        import_post_offices(conn, xlsx_file)
    else:
        print(f"Skipping railways/PO: not found in {data_dir} or Desktop")

    summary(conn)
    conn.close()

if __name__ == '__main__':
    main()
